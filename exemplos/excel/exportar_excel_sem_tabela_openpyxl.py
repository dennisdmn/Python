import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

saida = 'exemplo_saida.xlsx'
df = pd.DataFrame({'CONTA': ['1101', '1102'], 'VALOR': [100.0, -50.0]})

df.to_excel(saida, sheet_name='Resumo', index=False, startrow=5)

wb = load_workbook(saida)
ws = wb['Resumo']
ws.cell(row=1, column=1, value='Resumo')
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=6, column=col)
    cell.fill = PatternFill('solid', fgColor='1F3864')
    cell.font = Font(color='FFFFFF', bold=True)
ws.freeze_panes = 'A7'
wb.save(saida)
