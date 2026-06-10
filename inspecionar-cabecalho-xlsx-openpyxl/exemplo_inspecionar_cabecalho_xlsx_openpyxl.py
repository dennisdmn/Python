import openpyxl
from pathlib import Path

arquivo = Path(r"C:\Users\***\arquivo.xlsx")

wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)

print(f"Arquivo : {arquivo.name}")
print(f"Abas    : {wb.sheetnames}")

for nome_aba in wb.sheetnames:
    ws = wb[nome_aba]
    cabecalho = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    print(f"\nAba     : {nome_aba}")
    print(f"Colunas : {len(cabecalho)}")
    print(f"Linhas  : {ws.max_row - 1} (estimado)")  # max_row inclui o cabeçalho
    for i, col in enumerate(cabecalho, start=1):
        print(f"  {i:>3}. {col}")

wb.close()
