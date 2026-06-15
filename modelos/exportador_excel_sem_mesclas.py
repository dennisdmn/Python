import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


COR_AZUL = '1F3864'
COR_BRANCO = 'FFFFFF'


def exportar_excel_simples(caminho, abas):
    with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
        for nome, df in abas.items():
            df.to_excel(writer, sheet_name=nome, index=False, startrow=5)

    wb = load_workbook(caminho)
    for ws in wb.worksheets:
        ws.cell(row=1, column=1, value=ws.title)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=6, column=col)
            cell.fill = PatternFill('solid', fgColor=COR_AZUL)
            cell.font = Font(color=COR_BRANCO, bold=True)
        ws.freeze_panes = 'A7'
    wb.save(caminho)
