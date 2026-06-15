from pathlib import Path
from openpyxl import load_workbook


def ordenar_abas(caminho_excel: str | Path, ordem_desejada: list[str]) -> None:
    caminho = Path(caminho_excel)
    wb = load_workbook(caminho)
    abas_ordenadas = []
    for nome in ordem_desejada:
        if nome in wb.sheetnames:
            abas_ordenadas.append(wb[nome])
    abas_restantes = []
    for ws in wb.worksheets:
        if ws.title not in ordem_desejada:
            abas_restantes.append(ws)
    wb._sheets = abas_ordenadas + abas_restantes
    wb.save(caminho)
