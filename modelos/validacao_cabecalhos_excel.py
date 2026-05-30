"""Modelo simples para validar cabeçalhos de arquivos Excel.

Use quando um processo depende de colunas obrigatórias antes de ler uma base pesada.
Campos opcionais podem existir no layout, mas não bloqueiam a execução.
"""

from pathlib import Path

from openpyxl import load_workbook

ABA_PADRAO = "Exportação SAPUI5"
CAMPOS_CRITICOS = [
    "Conta do Razão",
    "Lançamento contábil",
    "Data de lançamento",
    "Mont.moeda empresa",
    "Empresa",
]
CAMPOS_OPCIONAIS = [
    "Chave de lançamento",
    "Nº de itens",
    "Divisão",
]


def ler_cabecalho_xlsx(caminho: Path, aba: str = ABA_PADRAO) -> list[str]:
    workbook = load_workbook(caminho, read_only=True, data_only=True)
    try:
        if aba not in workbook.sheetnames:
            raise ValueError(f"Aba '{aba}' não encontrada em {caminho.name}.")

        planilha = workbook[aba]
        primeira_linha = next(planilha.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return ["" if valor is None else str(valor).strip() for valor in primeira_linha]
    finally:
        workbook.close()


def validar_cabecalho(cabecalho: list[str]) -> tuple[str, list[str]]:
    campos_arquivo = set(cabecalho)
    faltando = [campo for campo in CAMPOS_CRITICOS if campo not in campos_arquivo]
    status = "OK" if not faltando else "REEXTRAIR"
    return status, faltando


if __name__ == "__main__":
    arquivo = Path(input("Informe o caminho do arquivo .xlsx: ").strip('"'))
    status, faltando = validar_cabecalho(ler_cabecalho_xlsx(arquivo))
    print(f"Status: {status}")
    if faltando:
        print("Campos faltando:", ", ".join(faltando))
