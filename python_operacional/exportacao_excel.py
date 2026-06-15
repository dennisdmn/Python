from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

COR_CABECALHO = '1F3864'
COR_BRANCO = 'FFFFFF'


def exportar_excel_seguro(caminho, abas, startrow=5):
    caminho = Path(caminho)
    caminho.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
        for nome_aba, df in abas.items():
            df.to_excel(writer, sheet_name=nome_aba, index=False, startrow=startrow)

    wb = load_workbook(caminho)
    for ws in wb.worksheets:
        ws.cell(row=1, column=1, value=ws.title)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=startrow + 1, column=col)
            cell.fill = PatternFill('solid', fgColor=COR_CABECALHO)
            cell.font = Font(color=COR_BRANCO, bold=True)
        ws.freeze_panes = 'A' + str(startrow + 2)
    wb.save(caminho)


def ordenar_abas(caminho, ordem):
    wb = load_workbook(caminho)
    principais = [wb[nome] for nome in ordem if nome in wb.sheetnames]
    restantes = [ws for ws in wb.worksheets if ws.title not in ordem]
    wb._sheets = principais + restantes
    wb.save(caminho)
