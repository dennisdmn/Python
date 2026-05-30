import argparse
import gc
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
import os
from pathlib import Path
import re
import sqlite3
import sys
from time import perf_counter, sleep
import unicodedata

from openpyxl import load_workbook
import polars as pl
import xlsxwriter

# Bibliotecas adotadas neste fluxo:

# - openpyxl: usado somente na etapa de validacao de cabecalhos. Apesar de ser mais lento
#   que ler o XML interno do XLSX diretamente, e uma biblioteca conhecida pela equipe,
#   simples de dar suporte e suficiente para uma etapa operacional previa.
#
# - polars: usado para ler cada XLSX e agrupar o movimento por conta antes de gravar no
#   SQLite. Assim, aproveitamos a performance do Polars sem manter todos os razoes em RAM.

# - xlsxwriter: usado para gerar o relatorio Excel de log com abas, tabelas e formatacao.
#   Gera arquivos .xlsx validos para o Microsoft Excel, evitando reparos no arquivo.

# - sqlite3: usado como camada temporaria na pasta de saida escolhida. Evita manter todos
#   os razoes em memoria e permite calcular os totais por conta de forma incremental.

# - argparse: usado para controlar a execucao por linha de comando, com modos dev,
#   teste e producao sem alterar constantes internas do arquivo.

# - dataclasses: usado para centralizar a configuracao de execucao em uma estrutura
#   simples, tipada e mais facil de manter do que parametros soltos.

# - gc, pathlib, datetime, os, re, sqlite3, sys, unicodedata, perf_counter e sleep:
#   bibliotecas padrao do Python para caminhos, timestamp, limpeza de textos,
#   banco temporario, abertura do relatorio no Windows, encoding do terminal,
#   medicao de tempo e retentativas de limpeza.

# Garante a exibicao correta de acentos no terminal do Windows.
sys.stdout.reconfigure(encoding="utf-8")

INICIO_SCRIPT = perf_counter()

# Caminhos e parametros principais.
# A pasta de entrada nao fica fixa aqui: ela deve ser selecionada pelo usuario
# ou informada por --pasta-bases para uso tecnico/automatizado.
TIMESTAMP_RELATORIO = datetime.now().strftime("%Y-%m-%d_%H%M")
FORMATO_NUMERO_CONTABIL = '_-#,##0.00_-;[Red](#,##0.00)_-;_-* "-"??_-;_-@_-'
EMPRESA_DESCONHECIDA = "empresa_nao_identificada"
PERIODO_DESCONHECIDO = "periodo_nao_identificado"
COR_TITULO_RELATORIO = "#5B9BD5"
COR_CABECALHO_TABELA = "#4F81BD"
COR_TOTAL_RELATORIO = "#FFFF99"
ESTILO_TABELA_PADRAO = "Table Style Medium 9"

ABA = "Exportação SAPUI5"

# O modo padrao de execucao e o "producao", adequado para uso final por duplo clique ou terminal.
# O modo "teste" limita a leitura para validar rapidamente o fluxo, e o modo "producao" tem postura mais restritiva
# O modo "dev" permanece disponivel para desenvolvimento e validacoes sem interromper por inconsistencia.
MODO_PADRAO = "producao"

# Agrupa os parametros efetivos da execucao em um contrato unico.
# Evita depender de varias constantes globais espalhadas pelo fluxo.
@dataclass(frozen=True)
class ConfiguracaoExecucao:
    modo: str
    pasta_bases: Path | None
    pasta_saida_resumo: Path | None
    arquivo_balancete: Path | None
    limite_arquivos: int | None
    escolher_pasta_entrada: bool
    escolher_pasta_saida: bool
    abrir_relatorio: bool
    parar_se_houver_inconsistencia: bool
    pausar_ao_final: bool

# Perfis de execucao:
# - dev: fluxo completo com janelas, sem interromper por inconsistencia.
# - teste: limita a leitura para validar rapidamente o script.
# - producao: postura mais restritiva para nao consolidar bases inconsistentes.
MODOS_EXECUCAO = {
    "dev": {
        "limite_arquivos": None,
        "escolher_pasta_entrada": True,
        "escolher_pasta_saida": True,
        "abrir_relatorio": False,
        "parar_se_houver_inconsistencia": False,
    },
    "teste": {
        "limite_arquivos": 1,
        "escolher_pasta_entrada": True,
        "escolher_pasta_saida": True,
        "abrir_relatorio": False,
        "parar_se_houver_inconsistencia": False,
    },
    "producao": {
        "limite_arquivos": None,
        "escolher_pasta_entrada": True,
        "escolher_pasta_saida": True,
        "abrir_relatorio": True,
        "parar_se_houver_inconsistencia": True,
    },
}

# Nomes tecnicos usados nas saidas intermediarias/finais.
# Os cabecalhos originais do SAP ficam concentrados aqui para facilitar manutencao.
COLUNA_ORIGEM = "arquivo_origem"
COLUNA_CONTA_RAZAO = "Conta do Razão"
COLUNA_DATA_LANCAMENTO = "Data de lançamento"
COLUNA_VALOR_MOVIMENTO = "Mont.moeda empresa"
COLUNA_CODIGO_CONTA_EXTRACAO = "codigo_conta_extracao"
COLUNA_CODIGO_CONTA = "codigo_conta"
COLUNA_DESCRICAO_CONTA = "descricao_conta"
COLUNA_MONTANTE_MOEDA_EMPRESA = "montante_moeda_empresa"
COLUNA_MONTANTE_AJUSTADO = "montante_ajustado"
COLUNA_SOMA_TRIMESTRAL = "soma_trimestral"
COLUNA_NOME_ARQUIVO = "nome_arquivo"
COLUNA_DATA_LANCTO_CONTABIL = "data_lancto_contabil"
COLUNA_CONTA_RAZAO_DIARIA = "conta_razao"
COLUNA_SALDOS_RAZOES = "saldos_razoes"
TABELA_RESUMO_ARQUIVO = "resumo_razoes_por_arquivo"
TABELA_RESUMO_DIA = "resumo_razoes_por_dia"
ABA_BALANCETE = "Sheet1"
LINHA_CABECALHO_BALANCETE = 27
PREFIXO_CODIGO_RAZAO_BALANCETE = "COG1/"
MARCADORES_TOTAL_BALANCETE = {"total", "soma total"}

CAMPOS_OBRIGATORIOS_BALANCETE = [
    "codigo_empresa",
    "descricao_empresa",
    "codigo_razao",
    "descricao_razao",
    "saldo_devedor_em_moeda_da_empresa",
    "saldo_credor_em_moeda_da_empresa",
]

# Lista minima esperada no layout SAP. A validacao permite colunas extras,
# mas exige somente estes nomes para manter a leitura posterior previsivel.
CAMPOS_OBRIGATORIOS_CRITICOS = [
    "Conta do Razão",
    "Lançamento contábil",
    "Data de lançamento",
    "Mont.moeda empresa",
    "Txt.it.partida indv.",
    "Centro de custo",
    "Empresa",
]

# Campos mantidos no layout interno do SAP, mas que nao bloqueiam a execucao
# porque nao sao usados nos calculos atuais da conciliacao.
CAMPOS_OBRIGATORIOS_OPCIONAIS = [
    "Chave de lançamento",
    "Nº de itens",
    "Divisão",
    "Tipo lçto.contábil",
]

def obter_pasta_bases(pasta_informada: Path | None, escolher_pasta_entrada: bool) -> Path:
    # A pasta de entrada deve ser escolhida pelo usuario ou informada explicitamente na CLI.
    if pasta_informada is not None:
        return pasta_informada

    # Sem pasta informada e sem janela nao ha entrada valida para processar.
    if not escolher_pasta_entrada:
        raise SystemExit("Informe --pasta-bases ou execute sem --sem-janela para selecionar a pasta das bases.")

    # Para usuarios novos, a janela inicia na pasta do usuario e nao em caminho local do desenvolvimento.
    pasta_inicial = Path.home()

    try:
        from tkinter import Tk, filedialog

        janela = Tk()
        janela.withdraw()
        janela.attributes("-topmost", True)

        pasta_escolhida = filedialog.askdirectory(
            title="Selecione a pasta com os arquivos de razao extraidos do SAP",
            initialdir=pasta_inicial,
        )
        janela.destroy()
    except Exception as exc:
        raise SystemExit(f"Nao foi possivel abrir a janela de selecao de pasta: {exc}") from exc

    if not pasta_escolhida:
        raise SystemExit("Processo cancelado: a pasta com os arquivos fonte deve ser selecionada.")

    return Path(pasta_escolhida)

def obter_pasta_saida_resumo(pasta_informada: Path | None, escolher_pasta_saida: bool) -> Path:
    # A pasta de saida deve ser escolhida pelo usuario ou informada explicitamente na CLI.
    if pasta_informada is not None:
        return pasta_informada

    if not escolher_pasta_saida:
        raise SystemExit("Informe --pasta-saida ou execute sem --sem-janela para selecionar a pasta de saida.")

    try:
        from tkinter import Tk, filedialog

        janela = Tk()
        janela.withdraw()
        janela.attributes("-topmost", True)

        pasta_escolhida = filedialog.askdirectory(
            title="Selecione a pasta para salvar o resumo trimestral e o log dos razoes",
            initialdir=Path.home(),
        )
        janela.destroy()
    except Exception as exc:
        raise SystemExit(f"Nao foi possivel abrir a janela de selecao da pasta de saida: {exc}") from exc

    if not pasta_escolhida:
        raise SystemExit("Processo cancelado: a pasta de saida deve ser selecionada.")

    return Path(pasta_escolhida)

def obter_arquivo_balancete(arquivo_informado: Path | None, escolher_arquivo: bool) -> Path:
    # O balancete e uma entrada obrigatoria da v6: via CLI ou janela de selecao de arquivo.
    if arquivo_informado is not None:
        arquivo_balancete = arquivo_informado
    else:
        if not escolher_arquivo:
            raise SystemExit("Informe --arquivo-balancete ou execute sem --sem-janela para selecionar o arquivo.")

        try:
            from tkinter import Tk, filedialog

            janela = Tk()
            janela.withdraw()
            janela.attributes("-topmost", True)

            arquivo_escolhido = filedialog.askopenfilename(
                title="Selecione o arquivo .xlsx do balancete",
                initialdir=Path.home(),
                filetypes=[("Arquivos Excel", "*.xlsx")],
            )
            janela.destroy()
        except Exception as exc:
            raise SystemExit(f"Nao foi possivel abrir a janela de selecao do balancete: {exc}") from exc

        if not arquivo_escolhido:
            raise SystemExit("Processo cancelado: o arquivo do balancete deve ser selecionado.")

        arquivo_balancete = Path(arquivo_escolhido)

    if not arquivo_balancete.exists() or not arquivo_balancete.is_file():
        raise SystemExit(f"Arquivo do balancete nao encontrado: {arquivo_balancete}")

    if arquivo_balancete.suffix.lower() != ".xlsx":
        raise SystemExit("O balancete deve ser um arquivo Excel no formato .xlsx.")

    return arquivo_balancete

def normalizar_nome_coluna(valor: object) -> str:
    # Converte cabecalhos livres do Excel para snake_case simples e sem acentos.
    texto = "" if valor is None else str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"[^a-z0-9]+", "_", texto).strip("_")
    return texto or "coluna_sem_nome"

def normalizar_cabecalhos_balancete(cabecalho_original: list[object]) -> list[str]:
    # As quatro primeiras colunas possuem nomes repetidos no arquivo fonte.
    cabecalhos: list[str] = []

    for indice, valor in enumerate(cabecalho_original):
        if indice == 0:
            nome = "codigo_empresa"
        elif indice == 1:
            nome = "descricao_empresa"
        elif indice == 2:
            nome = "codigo_razao"
        elif indice == 3:
            nome = "descricao_razao"
        else:
            nome = normalizar_nome_coluna(valor)

        nome_base = nome
        contador = 2
        while nome in cabecalhos:
            nome = f"{nome_base}_{contador}"
            contador += 1

        cabecalhos.append(nome)

    return cabecalhos

def validar_cabecalhos_balancete(cabecalhos: list[str]) -> None:
    # Validacao simples e sem log: falha cedo se o layout essencial mudar.
    campos_faltando = [campo for campo in CAMPOS_OBRIGATORIOS_BALANCETE if campo not in cabecalhos]
    if campos_faltando:
        raise SystemExit(f"Cabecalho do balancete invalido. Campos faltando: {', '.join(campos_faltando)}")

    if len(cabecalhos) != len(set(cabecalhos)):
        raise SystemExit("Cabecalho do balancete invalido: ha nomes duplicados apos a normalizacao.")

def linha_vazia(valores: tuple[object, ...]) -> bool:
    # Linhas totalmente vazias nao entram na aba final do balancete.
    return all(valor is None or str(valor).strip() == "" for valor in valores)


def texto_normalizado(valor: object) -> str:
    # Normaliza textos de controle para detectar totais sem depender do numero da linha.
    return "" if valor is None else str(valor).strip().lower()


def linha_total_balancete(valores: tuple[object, ...], cabecalhos: list[str]) -> bool:
    # Descarta linhas de total de forma dinamica, mesmo quando o balancete tiver outro tamanho.
    dados = dict(zip(cabecalhos, valores))
    codigo_razao = texto_normalizado(dados.get("codigo_razao"))
    descricao_razao = texto_normalizado(dados.get("descricao_razao"))
    codigo_empresa = texto_normalizado(dados.get("codigo_empresa"))

    if codigo_razao in MARCADORES_TOTAL_BALANCETE or descricao_razao in MARCADORES_TOTAL_BALANCETE:
        return True

    return codigo_empresa in MARCADORES_TOTAL_BALANCETE


def limpar_codigo_razao_balancete(valor: object) -> str:
    # Remove o prefixo operacional do SAP e mantem o codigo como texto para evitar perda de zeros.
    texto = "" if valor is None else str(valor).strip()
    if texto.startswith(PREFIXO_CODIGO_RAZAO_BALANCETE):
        return texto[len(PREFIXO_CODIGO_RAZAO_BALANCETE) :]

    return texto


def numero_balancete(valor: object) -> float:
    # Converte valores do balancete para calculo, aceitando numero nativo ou texto no padrao brasileiro.
    if valor is None or valor == "":
        return 0.0

    if isinstance(valor, int | float):
        return float(valor)

    texto = str(valor).strip().replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def arredondar_duas_casas(valor: object) -> float:
    # Replica o objetivo operacional do ARRED(valor; 2): grava valores ja fechados em 2 casas.
    numero = Decimal(str(numero_balancete(valor)))
    return float(numero.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

def formatar_data_periodo(valor: str) -> str:
    # Converte datas compactas do nome do balancete para leitura operacional no cabecalho.
    texto = str(valor).strip()
    if len(texto) == 8 and texto.isdigit():
        return f"{texto[:2]}/{texto[2:4]}/{texto[4:]}"
    return texto


def extrair_periodo_balancete(caminho_balancete: Path) -> str:
    # Periodo vem do nome padrao do arquivo: ...__ddmmaaaa__a__ddmmaaaa.xlsx.
    correspondencia = re.search(r"__(\d{8})__a__(\d{8})", caminho_balancete.stem)
    if not correspondencia:
        return PERIODO_DESCONHECIDO

    data_inicio, data_fim = correspondencia.groups()
    return f"{formatar_data_periodo(data_inicio)} a {formatar_data_periodo(data_fim)}"


def extrair_empresa_balancete(cabecalhos: list[str], linhas: list[list[object]]) -> str:
    # Usa a coluna normalizada do balancete para identificar a empresa no relatorio final.
    if "codigo_empresa" not in cabecalhos:
        return EMPRESA_DESCONHECIDA

    indice_empresa = cabecalhos.index("codigo_empresa")
    empresas = sorted({str(linha[indice_empresa]).strip() for linha in linhas if str(linha[indice_empresa]).strip()})
    if not empresas:
        return EMPRESA_DESCONHECIDA

    return "_".join(empresas)


def extrair_descricao_empresa_balancete(cabecalhos: list[str], linhas: list[list[object]]) -> str:
    # Recupera a descricao operacional da empresa para a linha 2 da conciliacao.
    if "descricao_empresa" not in cabecalhos:
        return "descricao_empresa_nao_identificada"

    indice_descricao = cabecalhos.index("descricao_empresa")
    descricoes = sorted({str(linha[indice_descricao]).strip() for linha in linhas if str(linha[indice_descricao]).strip()})
    if not descricoes:
        return "descricao_empresa_nao_identificada"

    return " | ".join(descricoes)


def montar_contexto_execucao(caminho_balancete: Path, cabecalhos: list[str], linhas: list[list[object]]) -> dict[str, str]:
    # Centraliza metadados usados no nome do arquivo e no cabecalho da conciliacao.
    return {
        "empresa": extrair_empresa_balancete(cabecalhos, linhas),
        "descricao_empresa": extrair_descricao_empresa_balancete(cabecalhos, linhas),
        "periodo": extrair_periodo_balancete(caminho_balancete),
        "elaborado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }


def nome_seguro_arquivo(valor: str) -> str:
    # Remove caracteres inadequados para nomes de arquivo mantendo identificadores legiveis.
    texto = re.sub(r"[^A-Za-z0-9_-]+", "_", str(valor).strip()).strip("_")
    return texto or EMPRESA_DESCONHECIDA

def inserir_movimento_periodo(cabecalhos: list[str], linha: list[object]) -> tuple[list[str], list[object]]:
    # Aplica a mesma regra dos razoes: contas que nao comecam com 1 tem o sinal invertido.
    if "movimento_periodo" not in cabecalhos:
        indice_insercao = cabecalhos.index("saldo_credor_em_moeda_da_empresa") + 1
        cabecalhos = cabecalhos[:indice_insercao] + ["movimento_periodo"] + cabecalhos[indice_insercao:]
    else:
        indice_insercao = cabecalhos.index("movimento_periodo")

    indice_codigo_razao = cabecalhos.index("codigo_razao")
    indice_devedor = cabecalhos.index("saldo_devedor_em_moeda_da_empresa")
    indice_credor = cabecalhos.index("saldo_credor_em_moeda_da_empresa")

    codigo_razao = str(linha[indice_codigo_razao]).strip()
    movimento_periodo = numero_balancete(linha[indice_devedor]) + numero_balancete(linha[indice_credor])
    if not codigo_razao.startswith("1"):
        movimento_periodo *= -1
    movimento_periodo = arredondar_duas_casas(movimento_periodo)

    linha = linha[:indice_insercao] + [movimento_periodo] + linha[indice_insercao:]
    return cabecalhos, linha

def ler_balancete(caminho_balancete: Path) -> tuple[list[str], list[list[object]]]:
    # Le somente o trecho util do balancete: cabecalho na linha 27 e dados antes dos totais.
    workbook = load_workbook(caminho_balancete, read_only=True, data_only=True)

    try:
        if ABA_BALANCETE not in workbook.sheetnames:
            raise SystemExit(f"Aba '{ABA_BALANCETE}' nao encontrada no balancete: {caminho_balancete.name}")

        planilha = workbook[ABA_BALANCETE]
        cabecalho_original = next(
            planilha.iter_rows(
                min_row=LINHA_CABECALHO_BALANCETE,
                max_row=LINHA_CABECALHO_BALANCETE,
                values_only=True,
            ),
            (),
        )
        cabecalhos = normalizar_cabecalhos_balancete(list(cabecalho_original))
        validar_cabecalhos_balancete(cabecalhos)
        cabecalhos_com_movimento: list[str] | None = None

        linhas: list[list[object]] = []
        for numero_linha, valores in enumerate(
            planilha.iter_rows(min_row=LINHA_CABECALHO_BALANCETE + 1, values_only=True),
            start=LINHA_CABECALHO_BALANCETE + 1,
        ):
            if linha_vazia(valores) or linha_total_balancete(valores, cabecalhos):
                continue

            linha = list(valores[: len(cabecalhos)])
            indice_codigo_razao = cabecalhos.index("codigo_razao")
            linha[indice_codigo_razao] = limpar_codigo_razao_balancete(linha[indice_codigo_razao])
            cabecalhos_com_movimento, linha = inserir_movimento_periodo(
                cabecalhos_com_movimento or cabecalhos,
                linha,
            )
            linhas.append(linha)

    finally:
        workbook.close()

    print(f"Balancete validado    : {caminho_balancete}")
    print(f"Linhas do balancete   : {len(linhas):,}")

    return cabecalhos_com_movimento or cabecalhos, linhas

def listar_arquivos_xlsx(pasta_bases: Path, limite_arquivos: int | None) -> list[Path]:
    # Processa somente arquivos Excel diretamente na pasta escolhida, sem subpastas.
    arquivos = sorted(pasta_bases.glob("*.xlsx"))

    # Limite usado em teste/dev para validar o fluxo sem ler todas as bases.
    if limite_arquivos is not None:
        arquivos = arquivos[:limite_arquivos]

    return arquivos

def ler_cabecalho_xlsx(caminho_arquivo: Path, nome_aba: str) -> list[str]:
    # Validação adotada: openpyxl, por ser mais conhecido pela equipe.
    workbook = load_workbook(caminho_arquivo, read_only=True, data_only=True)

    try:
        if nome_aba not in workbook.sheetnames:
            raise ValueError(f"Aba '{nome_aba}' nao encontrada. Abas disponiveis: {workbook.sheetnames}")

        planilha = workbook[nome_aba]
        # Somente a primeira linha e necessaria para validar o layout antes da carga pesada.
        primeira_linha = next(planilha.iter_rows(min_row=1, max_row=1, values_only=True), ())

        return ["" if valor is None else str(valor).strip() for valor in primeira_linha]
    finally:
        workbook.close()

def comparar_cabecalho(cabecalho: list[str]) -> tuple[str, list[str]]:
    # Ignora ordem, campos extras e opcionais; aponta somente campos criticos ausentes.
    # Usa set para comparar presenca dos campos sem depender da ordem das colunas.
    campos_arquivo = set(cabecalho)
    faltando = [campo for campo in CAMPOS_OBRIGATORIOS_CRITICOS if campo not in campos_arquivo]
    status = "OK" if not faltando else "REEXTRAIR"

    return status, faltando

def validar_cabecalhos(arquivos: list[Path], pasta_bases: Path) -> list[dict[str, str]]:
    # Percorre os arquivos e monta o mapa de status usado no log e no controle do fluxo.
    if not arquivos:
        print("Nenhum arquivo encontrado na pasta especificada.")
        return []

    inicio = perf_counter()
    # Cada item do mapeamento vira uma linha no relatorio de validacao.
    mapeamento = []

    print("=" * 80)
    print("ETAPA 1/2 - VALIDACAO DE CABECALHOS DOS RAZOES")
    print("=" * 80)
    print(f"Arquivos encontrados : {len(arquivos)}")
    print(f"Pasta das bases      : {pasta_bases}")
    print(f"Aba validada         : {ABA}")
    print(f"Campos criticos      : {len(CAMPOS_OBRIGATORIOS_CRITICOS)}")
    print()
    print("Progresso:")

    for indice, arquivo in enumerate(arquivos, start=1):
        try:
            cabecalho = ler_cabecalho_xlsx(arquivo, ABA)
            erro = ""
        except Exception as exc:
            cabecalho = []
            erro = str(exc)

        status, faltando = comparar_cabecalho(cabecalho)

        if erro:
            status = "ERRO"

        mapeamento.append(
            {
                "status": status,
                "arquivo": arquivo.name,
                "qtd_colunas": str(len(cabecalho)),
                "campos_faltando": "|".join(faltando),
                "erro": erro,
            }
        )

        percentual = indice / len(arquivos) * 100
        print(
            f"  {indice:>2}/{len(arquivos)} | "
            f"{percentual:6.2f}% | "
            f"{perf_counter() - inicio:6.2f}s | "
            f"{arquivo.name}"
        )

    return mapeamento

def contar_status(mapeamento: list[dict[str, str]]) -> tuple[int, int, int, int]:
    # Consolida os status para resumo em tela e no relatorio Excel.
    total = len(mapeamento)
    ok = sum(1 for item in mapeamento if item["status"] == "OK")
    reextrair = sum(1 for item in mapeamento if item["status"] == "REEXTRAIR")
    erro = sum(1 for item in mapeamento if item["status"] == "ERRO")

    return total, ok, reextrair, erro

def salvar_relatorio_xlsx(
    mapeamento: list[dict[str, str]],
    pasta_bases: Path,
    tempo_validacao: float,
    arquivo_relatorio: Path,
) -> None:
    # Gera o workbook de validacao com orientacao, resumo executivo e detalhe por arquivo.
    if not mapeamento:
        return

    arquivo_relatorio.parent.mkdir(parents=True, exist_ok=True)

    total, ok, reextrair, erro = contar_status(mapeamento)
    workbook = xlsxwriter.Workbook(arquivo_relatorio)

    fmt_titulo = workbook.add_format(
        {
            "bold": True,
            "font_size": 14,
            "font_color": "#111827",
            "bg_color": "#D9EAF7",
            "align": "left",
            "valign": "vcenter",
        }
    )
    fmt_rotulo = workbook.add_format({"bold": True, "font_color": "#111827", "valign": "vcenter"})
    fmt_ok = workbook.add_format({"bg_color": "#C6EFCE", "font_color": "#006100"})
    fmt_reextrair = workbook.add_format({"bg_color": "#FFEB9C", "font_color": "#9C6500"})
    fmt_erro = workbook.add_format({"bg_color": "#C00000", "font_color": "#FFFFFF"})
    fmt_secao = workbook.add_format(
        {"bold": True, "font_size": 12, "font_color": "#1F4E78", "bg_color": "#D9EAF7"}
    )
    fmt_texto = workbook.add_format({"text_wrap": True, "valign": "top"})
    fmt_alerta = workbook.add_format(
        {"text_wrap": True, "valign": "top", "bg_color": "#FFEB9C", "font_color": "#9C6500"}
    )

    # Aba textual: ajuda o usuario a interpretar status sem abrir o codigo.
    orientacao = workbook.add_worksheet("Orientacao")
    orientacao.hide_gridlines(2)
    orientacao.set_column("A:A", 28)
    orientacao.set_column("B:B", 105)
    orientacao.merge_range("A1:B1", "Orientacao ao usuario", fmt_titulo)
    orientacao.merge_range("A3:B3", "Objetivo da validacao", fmt_secao)
    orientacao.write("A4", "O que e validado", fmt_rotulo)
    orientacao.write(
        "B4",
        "Este arquivo verifica se cada razao extraida do SAP possui todos os campos criticos no cabecalho da aba "
        f"'{ABA}'. A validacao nao compara valores contabeis e nao valida a ordem das colunas.",
        fmt_texto,
    )
    orientacao.write("A5", "Campos criticos", fmt_rotulo)
    orientacao.write("B5", " | ".join(CAMPOS_OBRIGATORIOS_CRITICOS), fmt_texto)
    orientacao.merge_range("A7:B7", "Por que o layout padrao e necessario", fmt_secao)
    orientacao.write("A8", "Auditoria", fmt_rotulo)
    orientacao.write(
        "B8",
        "As bases serao utilizadas em conciliacao e enviadas para auditoria. O layout padrao reduz risco de perda de "
        "rastreabilidade, colunas incompletas e divergencia entre arquivos de periodos diferentes.",
        fmt_texto,
    )
    orientacao.merge_range("A10:B10", "Como interpretar o resultado", fmt_secao)
    orientacao.write("A11", "OK", fmt_rotulo)
    orientacao.write("B11", "O arquivo possui todos os campos criticos.", fmt_texto)
    orientacao.write("A12", "REEXTRAIR", fmt_rotulo)
    orientacao.write("B12", "O arquivo esta sem campos criticos. Extraia novamente a razao no SAP.", fmt_alerta)
    orientacao.write("A13", "ERRO", fmt_rotulo)
    orientacao.write("B13", "O arquivo nao pode ser lido tecnicamente.", fmt_texto)

    for linha in [4, 5, 8, 11, 12, 13]:
        orientacao.set_row(linha - 1, None, fmt_rotulo)

    # Aba executiva: totais de OK, REEXTRAIR e ERRO.
    resumo = workbook.add_worksheet("Resumo")
    resumo.hide_gridlines(2)
    resumo.set_column("A:A", 28)
    resumo.set_column("B:B", 72)
    resumo.merge_range("A1:B1", "Resumo da Validacao de Cabecalhos dos Razoes", fmt_titulo)
    resumo.set_row(0, 26)
    resumo.write("A3", "Data/Hora do relatorio", fmt_rotulo)
    resumo.write("B3", datetime.now().strftime("%Y-%m-%d %H:%M"))
    resumo.write("A4", "Pasta das bases", fmt_rotulo)
    resumo.write("B4", str(pasta_bases))
    resumo.write("A5", "Aba validada", fmt_rotulo)
    resumo.write("B5", ABA)
    resumo.write("A6", "Campos criticos", fmt_rotulo)
    resumo.write("B6", len(CAMPOS_OBRIGATORIOS_CRITICOS))

    resumo_dados = [
        ["Arquivos validados", total],
        ["Arquivos OK", ok],
        ["Reextrair do SAP", reextrair],
        ["Erros de leitura", erro],
        ["Tempo validacao (segundos)", round(tempo_validacao, 2)],
    ]
    resumo.add_table(
        7,
        0,
        12,
        1,
        {
            "name": "TabelaResumo",
            "style": "Table Style Medium 2",
            "columns": [{"header": "Indicador"}, {"header": "Valor"}],
            "data": resumo_dados,
        },
    )
    resumo.write("B10", ok, fmt_ok)
    resumo.write("B11", reextrair, fmt_reextrair)
    resumo.write("B12", erro, fmt_erro if erro else fmt_ok)

    # Aba detalhada: uma linha por arquivo analisado.
    validacao = workbook.add_worksheet("Validacao")
    validacao.freeze_panes(1, 0)
    validacao.set_column("A:A", 14)
    validacao.set_column("B:B", 46)
    validacao.set_column("C:C", 12)
    validacao.set_column("D:D", 46)
    validacao.set_column("E:E", 36)

    colunas = ["status", "arquivo", "qtd_colunas", "campos_faltando", "erro"]
    dados = [[item[coluna] for coluna in colunas] for item in mapeamento]
    validacao.add_table(
        0,
        0,
        len(dados),
        len(colunas) - 1,
        {
            "name": "TabelaValidacao",
            "style": "Table Style Medium 2",
            "columns": [{"header": coluna} for coluna in colunas],
            "data": dados,
        },
    )

    for linha, item in enumerate(mapeamento, start=1):
        formato = {"OK": fmt_ok, "REEXTRAIR": fmt_reextrair, "ERRO": fmt_erro}.get(item["status"])
        if formato is not None:
            validacao.write(linha, 0, item["status"], formato)

    workbook.close()

def abrir_relatorio_se_necessario(reextrair: int, erro: int, abrir_relatorio: bool, arquivo_relatorio: Path) -> None:
    # Abre o Excel automaticamente apenas quando configurado e quando houver algo a revisar.
    if not abrir_relatorio or (reextrair == 0 and erro == 0):
        return

    try:
        print()
        print("Inconsistencias encontradas. Abrindo relatorio Excel...")
        os.startfile(arquivo_relatorio)
    except OSError as exc:
        print(f"Nao foi possivel abrir o relatorio automaticamente: {exc}")

def abrir_excel_final_se_necessario(abrir_relatorio: bool, arquivo_resumo: Path) -> None:
    # Em producao, abre o arquivo final automaticamente quando a conciliacao termina sem bloqueios.
    if not abrir_relatorio:
        return

    try:
        print()
        print("Conciliacao finalizada. Abrindo arquivo Excel final...")
        os.startfile(arquivo_resumo)
    except OSError as exc:
        print(f"Nao foi possivel abrir o arquivo Excel final automaticamente: {exc}")

def criar_sqlite_temporario(pasta_trabalho: Path) -> Path:
    # Prepara a estrutura minima do banco intermediario usado na consolidacao incremental.
    # O banco fica em subpasta operacional da saida e e apagado ao final do processamento.
    pasta_trabalho.mkdir(parents=True, exist_ok=True)
    caminho_sqlite = pasta_trabalho / f"{TIMESTAMP_RELATORIO}_razoes_temp.sqlite"

    # Defesa contra colisao improvavel de nome no diretorio temporario.
    if caminho_sqlite.exists():
        caminho_sqlite.unlink()

    with sqlite3.connect(caminho_sqlite) as conexao:
        # Guarda somente o resumo por arquivo/conta, nao os lancamentos completos.
        conexao.execute(
            f"""
            CREATE TABLE {TABELA_RESUMO_ARQUIVO} (
                {COLUNA_ORIGEM} TEXT NOT NULL,
                {COLUNA_CODIGO_CONTA_EXTRACAO} TEXT NOT NULL,
                {COLUNA_CODIGO_CONTA} TEXT NOT NULL,
                {COLUNA_DESCRICAO_CONTA} TEXT NOT NULL,
                {COLUNA_MONTANTE_MOEDA_EMPRESA} REAL NOT NULL
            )
            """
        )
        conexao.execute(
            f"""
            -- Indice simples para acelerar a soma final por conta.
            CREATE INDEX idx_{TABELA_RESUMO_ARQUIVO}_conta
            ON {TABELA_RESUMO_ARQUIVO} ({COLUNA_CODIGO_CONTA})
            """
        )
        conexao.execute(
            f"""
            CREATE TABLE {TABELA_RESUMO_DIA} (
                {COLUNA_NOME_ARQUIVO} TEXT NOT NULL,
                {COLUNA_DATA_LANCTO_CONTABIL} TEXT NOT NULL,
                {COLUNA_CONTA_RAZAO_DIARIA} TEXT NOT NULL,
                {COLUNA_SALDOS_RAZOES} REAL NOT NULL
            )
            """
        )
        conexao.execute(
            f"""
            -- Indice simples para acelerar a leitura da evidencia diaria.
            CREATE INDEX idx_{TABELA_RESUMO_DIA}_arquivo_data_conta
            ON {TABELA_RESUMO_DIA} ({COLUNA_NOME_ARQUIVO}, {COLUNA_DATA_LANCTO_CONTABIL}, {COLUNA_CONTA_RAZAO_DIARIA})
            """
        )

    return caminho_sqlite

def apagar_sqlite_temporario(caminho_sqlite: Path, tentativas: int = 10, intervalo: float = 0.5) -> bool:
    # Remove o banco temporario com retentativas para contornar locks breves do Windows.
    arquivos_sqlite = [
        caminho_sqlite,
        caminho_sqlite.with_name(f"{caminho_sqlite.name}-journal"),
        caminho_sqlite.with_name(f"{caminho_sqlite.name}-wal"),
        caminho_sqlite.with_name(f"{caminho_sqlite.name}-shm"),
    ]

    for tentativa in range(1, tentativas + 1):
        gc.collect()

        for arquivo in arquivos_sqlite:
            try:
                arquivo.unlink(missing_ok=True)
            except PermissionError:
                pass

        if not any(arquivo.exists() for arquivo in arquivos_sqlite):
            print()
            print("SQLite temporario apagado apos a execucao.")
            return True

        if tentativa < tentativas:
            sleep(intervalo)

    print()
    print("Aviso: nao foi possivel apagar o SQLite temporario porque o arquivo ainda esta em uso.")
    print("Feche processos que possam estar lendo o arquivo e remova manualmente, se ele permanecer:")
    for arquivo in arquivos_sqlite:
        if arquivo.exists():
            print(f"  {arquivo}")

    return False

def carregar_resumos_no_sqlite(arquivos: list[Path], caminho_sqlite: Path) -> None:
    # Lê um XLSX por vez, resume, grava no SQLite e descarta o DataFrame grande.
    if not arquivos:
        print("Nenhum arquivo encontrado para leitura com Polars.")
        return

    total_arquivos = len(arquivos)
    inicio_total = perf_counter()
    total_linhas_lidas = 0

    print()
    print("=" * 80)
    print("ETAPA 2/2 - LEITURA INCREMENTAL E AGRUPAMENTO DOS RAZOES")
    print("=" * 80)
    print(f"Arquivos para leitura : {total_arquivos}")
    print(f"Aba lida              : {ABA}")
    print(f"SQLite temporario     : {caminho_sqlite}")
    print()

    with sqlite3.connect(caminho_sqlite) as conexao:
        for indice, arquivo in enumerate(arquivos, start=1):
            # DataFrame temporario: existe somente durante o processamento deste arquivo.
            # Leitura completa do arquivo corrente; o DataFrame e descartado ao fim da iteracao.
            df = pl.read_excel(
                arquivo,
                sheet_name=ABA,
                engine="calamine",
                infer_schema_length=None,
            )
            total_linhas_lidas += df.height

            # Conversao robusta do valor: aceita numero nativo e texto no padrao brasileiro.
            valor_numerico = pl.col(COLUNA_VALOR_MOVIMENTO).cast(pl.Float64, strict=False)
            valor_texto_br = (
                pl.col(COLUNA_VALOR_MOVIMENTO)
                .cast(pl.Utf8)
                .str.replace_all(r"\.", "")
                .str.replace(",", ".")
                .cast(pl.Float64, strict=False)
            )

            # Regra de validacao contábil:
            # contas iniciadas por 1 mantem o sinal original; demais grupos contabeis
            # sao invertidos para que o resumo fique no sentido esperado pelo usuario.
            montante_ajustado = (
                pl.when(pl.col(COLUNA_CODIGO_CONTA).str.starts_with("1"))
                .then(pl.col(COLUNA_MONTANTE_MOEDA_EMPRESA))
                .otherwise(pl.col(COLUNA_MONTANTE_MOEDA_EMPRESA) * -1)
                .alias(COLUNA_MONTANTE_AJUSTADO)
            )

            # Base ja higienizada: alimenta tanto o resumo por arquivo quanto a evidencia diaria.
            data_lancto_formatada = (
                pl.col(COLUNA_DATA_LANCAMENTO)
                .cast(pl.Date, strict=False)
                .dt.strftime("%d/%m/%Y")
                .fill_null(pl.col(COLUNA_DATA_LANCAMENTO).cast(pl.Utf8))
                .alias(COLUNA_DATA_LANCTO_CONTABIL)
            )
            df_base_arquivo = (
                df.select([COLUNA_CONTA_RAZAO, COLUNA_DATA_LANCAMENTO, COLUNA_VALOR_MOVIMENTO])
                .with_columns(
                    pl.col(COLUNA_CONTA_RAZAO).cast(pl.Utf8).str.strip_chars().alias(COLUNA_CODIGO_CONTA_EXTRACAO),
                    pl.coalesce([valor_numerico, valor_texto_br]).alias(COLUNA_MONTANTE_MOEDA_EMPRESA),
                    data_lancto_formatada,
                )
                .with_columns(
                    pl.col(COLUNA_CODIGO_CONTA_EXTRACAO).str.extract(r"^\s*([^\s(]+)", 1).alias(COLUNA_CODIGO_CONTA),
                    # Captura a descricao pelo parenteses externo para preservar marcadores internos, como (-).
                    pl.col(COLUNA_CODIGO_CONTA_EXTRACAO)
                    .str.extract(r"^\s*[^\s(]+\s+\((.*)\)\s*$", 1)
                    .fill_null("")
                    .alias(COLUNA_DESCRICAO_CONTA),
                )
                # Descarta linhas sem conta extraida para evitar agrupamentos sem chave contabil.
                .filter(pl.col(COLUNA_CODIGO_CONTA).is_not_null() & (pl.col(COLUNA_CODIGO_CONTA) != ""))
                .with_columns(montante_ajustado)
            )

            # Reduz o arquivo ao menor conjunto necessario antes de persistir no SQLite.
            df_resumo_arquivo = (
                df_base_arquivo
                .group_by([COLUNA_CODIGO_CONTA_EXTRACAO, COLUNA_CODIGO_CONTA, COLUNA_DESCRICAO_CONTA])
                .agg(pl.col(COLUNA_MONTANTE_AJUSTADO).sum().fill_null(0).alias(COLUNA_MONTANTE_MOEDA_EMPRESA))
                .with_columns(pl.lit(arquivo.name).alias(COLUNA_ORIGEM))
                .select(
                    [
                        COLUNA_ORIGEM,
                        COLUNA_CODIGO_CONTA_EXTRACAO,
                        COLUNA_CODIGO_CONTA,
                        COLUNA_DESCRICAO_CONTA,
                        COLUNA_MONTANTE_MOEDA_EMPRESA,
                    ]
                )
            )

            # Evidencia diaria: ajuda a localizar o arquivo/dia/conta que precisa ser substituido.
            df_resumo_dia = (
                df_base_arquivo
                .group_by([COLUNA_DATA_LANCTO_CONTABIL, COLUNA_CODIGO_CONTA])
                .agg(pl.col(COLUNA_MONTANTE_AJUSTADO).sum().fill_null(0).alias(COLUNA_SALDOS_RAZOES))
                .with_columns(pl.lit(arquivo.name).alias(COLUNA_NOME_ARQUIVO))
                .select(
                    [
                        COLUNA_NOME_ARQUIVO,
                        COLUNA_DATA_LANCTO_CONTABIL,
                        pl.col(COLUNA_CODIGO_CONTA).alias(COLUNA_CONTA_RAZAO_DIARIA),
                        COLUNA_SALDOS_RAZOES,
                    ]
                )
            )

            # Insercao incremental: o SQLite recebe apenas poucos registros por conta/arquivo.
            conexao.executemany(
                f"""
                INSERT INTO {TABELA_RESUMO_ARQUIVO}
                (
                    {COLUNA_ORIGEM},
                    {COLUNA_CODIGO_CONTA_EXTRACAO},
                    {COLUNA_CODIGO_CONTA},
                    {COLUNA_DESCRICAO_CONTA},
                    {COLUNA_MONTANTE_MOEDA_EMPRESA}
                )
                VALUES (?, ?, ?, ?, ?)
                """,
                df_resumo_arquivo.rows(),
            )
            conexao.executemany(
                f"""
                INSERT INTO {TABELA_RESUMO_DIA}
                (
                    {COLUNA_NOME_ARQUIVO},
                    {COLUNA_DATA_LANCTO_CONTABIL},
                    {COLUNA_CONTA_RAZAO_DIARIA},
                    {COLUNA_SALDOS_RAZOES}
                )
                VALUES (?, ?, ?, ?)
                """,
                df_resumo_dia.rows(),
            )
            # Commit por arquivo reduz perda de progresso se um arquivo posterior falhar.
            conexao.commit()

            percentual = indice / total_arquivos * 100
            print(
                f"  {indice:>2}/{total_arquivos} | "
                f"{percentual:6.2f}% | "
                f"{perf_counter() - inicio_total:6.2f}s | "
                f"{arquivo.name}"
            )

    tempo_total = perf_counter() - inicio_total

    print("\n" + "=" * 80)
    print("RESUMO DA LEITURA INCREMENTAL")
    print("=" * 80)
    print(f"Arquivos lidos       : {total_arquivos}")
    print(f"Linhas lidas         : {total_linhas_lidas:,}")
    print(f"Tempo leitura/agrup. : {tempo_total:.2f} segundos")
    print("=" * 80)

def criar_formatos_relatorio(workbook: xlsxwriter.Workbook) -> dict[str, object]:
    # Formatos compartilhados para manter o mesmo layout em todas as abas do arquivo final.
    return {
        "titulo": workbook.add_format(
            {
                "bold": True,
                "font_size": 14,
                "font_color": "#FFFFFF",
                "bg_color": COR_TITULO_RELATORIO,
                "align": "left",
                "valign": "vcenter",
            }
        ),
        "subtitulo": workbook.add_format({"font_color": "#4B5563", "italic": True}),
        "menu_texto": workbook.add_format({"text_wrap": True, "valign": "top", "border": 1}),
        "menu_ordem": workbook.add_format({"align": "center", "valign": "top", "border": 1}),
        "menu_destaque_vermelho": workbook.add_format({"font_color": "#FF0000", "bold": True}),
        "link_menu_retorno": workbook.add_format({"font_color": "#0000FF", "underline": True, "align": "center"}),
        "menu_link": workbook.add_format(
            {
                "font_color": "#0563C1",
                "underline": True,
                "bold": True,
                "valign": "top",
                "border": 1,
            }
        ),
        "numero": workbook.add_format({"num_format": FORMATO_NUMERO_CONTABIL}),
        "cabecalho_tabela": workbook.add_format(
            {
                "bold": True,
                "font_color": "#FFFFFF",
                "bg_color": COR_CABECALHO_TABELA,
                "border": 1,
            }
        ),
        "total_rotulo": workbook.add_format(
            {
                "bold": True,
                "bg_color": COR_TOTAL_RELATORIO,
                "font_color": "#000000",
                "align": "left",
                "valign": "vcenter",
            }
        ),
        "total_numero": workbook.add_format(
            {
                "bold": True,
                "bg_color": COR_TOTAL_RELATORIO,
                "num_format": FORMATO_NUMERO_CONTABIL,
            }
        ),
    }


def escrever_cabecalho_relatorio(
    planilha: xlsxwriter.worksheet.Worksheet,
    formatos: dict[str, object],
    ultima_coluna: int,
    titulo: str,
    contexto: dict[str, str],
    mostrar_link_menu: bool = False,
    coluna_link_menu: int | None = None,
) -> None:
    # Cabecalho padrao das abas finais: titulo, empresa e data/hora de elaboracao.
    planilha.merge_range(0, 0, 0, ultima_coluna, titulo, formatos["titulo"])
    planilha.write("A2", f"Empresa: {contexto['descricao_empresa']}", formatos["subtitulo"])
    planilha.write("A3", f"Elaborado em: {contexto['elaborado_em']}", formatos["subtitulo"])
    if mostrar_link_menu:
        coluna_destino_menu = coluna_link_menu if coluna_link_menu is not None else ultima_coluna + 1
        planilha.write_url(1, coluna_destino_menu, "internal:'Menu'!A1", formatos["link_menu_retorno"], string="Menu")
    planilha.set_row(0, 24)


def colunas_tabela_padrao(colunas: list[str], formatos: dict[str, object]) -> list[dict[str, object]]:
    # Garante letras brancas e fundo azul nos cabecalhos das tabelas do Excel.
    return [{"header": coluna, "header_format": formatos["cabecalho_tabela"]} for coluna in colunas]


def escrever_aba_menu(workbook: xlsxwriter.Workbook, contexto: dict[str, str]) -> None:
    # Aba inicial do workbook: orienta a navegacao e explica o objetivo de cada evidencia.
    planilha = workbook.add_worksheet("Menu")
    planilha.hide_gridlines(2)
    planilha.set_tab_color(COR_TITULO_RELATORIO)

    formatos = criar_formatos_relatorio(workbook)
    escrever_cabecalho_relatorio(planilha, formatos, 3, "Menu do Relatorio de Conciliacao Contabil", contexto)

    colunas = ["ordem", "aba", "objetivo", "uso_recomendado"]
    itens_menu = [
        (
            1,
            "Conciliacao",
            "Compara o movimento do balancete contra os razoes agrupados por conta contabil.",
            "Principal aba de analise; use para identificar diferencas entre SAP HANA e razoes SAP.",
        ),
        (
            2,
            "Balancete",
            "Exibe o balancete na integra do SAP HANA, normalizado apenas para padronizar cabecalhos, sinais e formatos.",
            "Use para conferir a base oficial de comparacao, empresa, contas, saldos e movimento do periodo.",
        ),
        (
            3,
            "Resumo - Razões",
            "Mostra os razoes consolidados por conta contabil no periodo processado.",
            "Use para validar o total agrupado dos arquivos de razao extraidos do SAP.",
        ),
        (
            4,
            "Razoes por Dia",
            "Detalha os saldos dos razoes por arquivo, data de lancamento e conta contabil.",
            "Use para investigar qual arquivo ou dia explica uma divergencia encontrada na conciliacao.",
        ),
    ]

    planilha.set_column("A:A", 10)
    planilha.set_column("B:B", 24)
    planilha.set_column("C:C", 50)
    planilha.set_column("D:D", 50)
    planilha.set_row(4, 22)

    for indice_coluna, coluna in enumerate(colunas):
        planilha.write(4, indice_coluna, coluna, formatos["cabecalho_tabela"])

    for indice_linha, (ordem, aba, objetivo, uso_recomendado) in enumerate(itens_menu, start=5):
        planilha.write_number(indice_linha, 0, ordem, formatos["menu_ordem"])
        planilha.write_url(indice_linha, 1, f"internal:'{aba}'!A1", formatos["menu_link"], string=aba)
        if aba == "Balancete":
            planilha.write_rich_string(
                indice_linha,
                2,
                "Exibe o ",
                formatos["menu_destaque_vermelho"],
                "balancete na integra do SAP HANA",
                ", normalizado apenas para padronizar cabecalhos, sinais e formatos.",
                formatos["menu_texto"],
            )
        else:
            planilha.write(indice_linha, 2, objetivo, formatos["menu_texto"])
        planilha.write(indice_linha, 3, uso_recomendado, formatos["menu_texto"])
        planilha.set_row(indice_linha, 29)


def escrever_aba_razoes_por_dia(
    workbook: xlsxwriter.Workbook,
    linhas_dia: list[tuple[object, object, object, object]],
    contexto: dict[str, str],
) -> None:
    # Evidencia operacional para localizar arquivo, dia e conta com possivel necessidade de substituicao.
    planilha = workbook.add_worksheet("Razoes por Dia")
    planilha.hide_gridlines(2)
    planilha.freeze_panes(5, 0)
    planilha.set_tab_color(COR_TITULO_RELATORIO)

    colunas = [
        COLUNA_NOME_ARQUIVO,
        COLUNA_DATA_LANCTO_CONTABIL,
        COLUNA_CONTA_RAZAO_DIARIA,
        COLUNA_SALDOS_RAZOES,
    ]
    limite_linhas_excel = 1_048_571
    if len(linhas_dia) > limite_linhas_excel:
        raise SystemExit(
            "A evidencia diaria ultrapassa o limite de linhas do Excel. "
            "Avalie filtrar por periodo/arquivo ou dividir a saida."
        )

    formatos = criar_formatos_relatorio(workbook)
    escrever_cabecalho_relatorio(planilha, formatos, len(colunas) - 1, "Saldos dos Razoes por Arquivo, Dia e Conta", contexto, True, 3)

    # Linha de totais: soma apenas o saldo dos razoes por dia, mantendo as demais colunas como contexto textual.
    total_saldos_razoes = arredondar_duas_casas(sum(numero_balancete(linha[3]) for linha in linhas_dia))
    planilha.write("A4", "TOTAIS:", formatos["total_rotulo"])
    planilha.write_blank("B4", None, formatos["total_rotulo"])
    planilha.write_blank("C4", None, formatos["total_rotulo"])
    planilha.write_number("D4", total_saldos_razoes, formatos["total_numero"])
    planilha.set_row(3, 20)
    planilha.set_column("A:A", 46)
    planilha.set_column("B:B", 18)
    planilha.set_column("C:C", 18)
    planilha.set_column("D:D", 20, formatos["numero"])

    linha_cabecalho = 4
    if linhas_dia:
        planilha.add_table(
            linha_cabecalho,
            0,
            linha_cabecalho + len(linhas_dia),
            len(colunas) - 1,
            {
                "name": "TabelaRazoesPorDia",
                "style": ESTILO_TABELA_PADRAO,
                "columns": colunas_tabela_padrao(colunas, formatos),
                "data": linhas_dia,
            },
        )
    else:
        for indice_coluna, coluna in enumerate(colunas):
            planilha.write(linha_cabecalho, indice_coluna, coluna, formatos["cabecalho_tabela"])

def chave_ordenacao_codigo(codigo: str) -> tuple[int, int | str]:
    # Ordena contas numericas pela grandeza do codigo e deixa codigos nao numericos ao final.
    texto = str(codigo).strip()
    return (0, int(texto)) if texto.isdigit() else (1, texto)


def montar_linhas_conciliacao(
    linhas_resumo: list[tuple[object, object, object, object]],
    cabecalhos_balancete: list[str],
    linhas_balancete: list[list[object]],
) -> list[list[object]]:
    # Equivale ao UNION DISTINCT dos codigos do balancete com os codigos dos razoes agrupados.
    indice_codigo_balancete = cabecalhos_balancete.index("codigo_razao")
    indice_descricao_balancete = cabecalhos_balancete.index("descricao_razao")
    indice_movimento_balancete = cabecalhos_balancete.index("movimento_periodo")

    descricoes_balancete: dict[str, str] = {}
    montantes_balancete: dict[str, float] = {}
    for linha in linhas_balancete:
        codigo = str(linha[indice_codigo_balancete]).strip()
        if not codigo:
            continue

        descricoes_balancete.setdefault(codigo, str(linha[indice_descricao_balancete]).strip())
        montantes_balancete[codigo] = montantes_balancete.get(codigo, 0.0) + numero_balancete(
            linha[indice_movimento_balancete]
        )

    descricoes_razoes: dict[str, str] = {}
    montantes_razoes: dict[str, float] = {}
    for _, codigo_conta, descricao_conta, soma_trimestral in linhas_resumo:
        codigo = str(codigo_conta).strip()
        if not codigo:
            continue

        # Prioridade de descricao: razoes agrupados; balancete fica como fallback.
        descricoes_razoes.setdefault(codigo, str(descricao_conta).strip())
        montantes_razoes[codigo] = montantes_razoes.get(codigo, 0.0) + numero_balancete(soma_trimestral)

    codigos = sorted(set(montantes_balancete) | set(montantes_razoes), key=chave_ordenacao_codigo)
    linhas_conciliacao: list[list[object]] = []
    for codigo in codigos:
        montante_balancete = arredondar_duas_casas(montantes_balancete.get(codigo, 0.0))
        montante_razoes = arredondar_duas_casas(montantes_razoes.get(codigo, 0.0))
        conf_montantes = arredondar_duas_casas(montante_balancete - montante_razoes)
        linhas_conciliacao.append(
            [
                codigo,
                descricoes_razoes.get(codigo) or descricoes_balancete.get(codigo, ""),
                montante_balancete,
                montante_razoes,
                conf_montantes,
            ]
        )

    return linhas_conciliacao


def escrever_aba_conciliacao(
    workbook: xlsxwriter.Workbook,
    linhas_resumo: list[tuple[object, object, object, object]],
    cabecalhos_balancete: list[str],
    linhas_balancete: list[list[object]],
    contexto: dict[str, str],
) -> None:
    # Primeira aba do arquivo: confronto objetivo entre balancete e razoes agrupados por conta.
    planilha = workbook.add_worksheet("Conciliacao")
    planilha.hide_gridlines(2)
    planilha.freeze_panes(5, 0)
    planilha.set_tab_color(COR_TITULO_RELATORIO)

    colunas = [
        "codigo_conta_contabil",
        "descricao_conta_contabil",
        "montante_balancete",
        "montante_razoes",
        "conf_montantes",
    ]
    linhas_conciliacao = montar_linhas_conciliacao(linhas_resumo, cabecalhos_balancete, linhas_balancete)
    total_balancete = arredondar_duas_casas(sum(numero_balancete(linha[2]) for linha in linhas_conciliacao))
    total_razoes = arredondar_duas_casas(sum(numero_balancete(linha[3]) for linha in linhas_conciliacao))
    total_conferencia = arredondar_duas_casas(sum(numero_balancete(linha[4]) for linha in linhas_conciliacao))

    formatos = criar_formatos_relatorio(workbook)
    fmt_numero = formatos["numero"]
    fmt_ok = workbook.add_format({"font_color": "#166534", "num_format": FORMATO_NUMERO_CONTABIL})
    fmt_divergencia = workbook.add_format({"font_color": "#991B1B", "bold": True, "num_format": FORMATO_NUMERO_CONTABIL})

    titulo_conciliacao = f"Conciliacao Balancete x Razoes | Empresa {contexto['empresa']} | Periodo {contexto['periodo']}"
    escrever_cabecalho_relatorio(planilha, formatos, len(colunas) - 1, titulo_conciliacao, contexto, True, 4)
    planilha.write("A4", "TOTAIS:", formatos["total_rotulo"])
    planilha.write_blank("B4", None, formatos["total_rotulo"])
    planilha.write_number("C4", total_balancete, formatos["total_numero"])
    planilha.write_number("D4", total_razoes, formatos["total_numero"])
    planilha.write_number("E4", total_conferencia, formatos["total_numero"])
    planilha.set_row(0, 24)
    planilha.set_row(3, 20)

    planilha.set_column("A:A", 24)
    planilha.set_column("B:B", 40)
    planilha.set_column("C:E", 20, fmt_numero)

    linha_cabecalho = 4
    if linhas_conciliacao:
        planilha.add_table(
            linha_cabecalho,
            0,
            linha_cabecalho + len(linhas_conciliacao),
            len(colunas) - 1,
            {
                "name": "TabelaConciliacao",
                "style": ESTILO_TABELA_PADRAO,
                "columns": colunas_tabela_padrao(colunas, formatos),
                "data": linhas_conciliacao,
            },
        )
    else:
        for indice_coluna, coluna in enumerate(colunas):
            planilha.write(linha_cabecalho, indice_coluna, coluna, formatos["cabecalho_tabela"])

    primeira_linha_dados = linha_cabecalho + 1
    ultima_linha_dados = linha_cabecalho + max(len(linhas_conciliacao), 1)
    planilha.conditional_format(primeira_linha_dados, 4, ultima_linha_dados, 4, {"type": "cell", "criteria": "=", "value": 0, "format": fmt_ok})
    planilha.conditional_format(primeira_linha_dados, 4, ultima_linha_dados, 4, {"type": "cell", "criteria": "!=", "value": 0, "format": fmt_divergencia})


def escrever_aba_balancete(
    workbook: xlsxwriter.Workbook,
    cabecalhos: list[str],
    linhas: list[list[object]],
    contexto: dict[str, str],
) -> None:
    # Inclui o balancete normalizado no mesmo arquivo final, sem gerar log separado.
    planilha = workbook.add_worksheet("Balancete")
    planilha.hide_gridlines(2)
    planilha.set_zoom(85)
    planilha.freeze_panes(5, 0)
    planilha.set_tab_color(COR_TITULO_RELATORIO)

    ultima_coluna = max(len(cabecalhos) - 1, 0)
    formatos = criar_formatos_relatorio(workbook)
    escrever_cabecalho_relatorio(planilha, formatos, ultima_coluna, "Balancete Normalizado", contexto, True, 8)

    # Linha de totais do balancete: somente colunas monetarias sao somadas.
    colunas_total_balancete = {
        "saldo_inicial_em_moeda_da_empresa",
        "saldo_devedor_em_moeda_da_empresa",
        "saldo_credor_em_moeda_da_empresa",
        "movimento_periodo",
        "saldo_final_na_moeda_da_empresa",
    }
    totais_balancete: dict[str, float] = {}
    for indice_coluna, cabecalho in enumerate(cabecalhos):
        if cabecalho in colunas_total_balancete:
            totais_balancete[cabecalho] = arredondar_duas_casas(
                sum(numero_balancete(linha[indice_coluna]) for linha in linhas)
            )

    planilha.write("A4", "TOTAIS:", formatos["total_rotulo"])
    for indice_coluna, cabecalho in enumerate(cabecalhos[1:], start=1):
        if cabecalho in totais_balancete:
            planilha.write_number(3, indice_coluna, totais_balancete[cabecalho], formatos["total_numero"])
        else:
            planilha.write_blank(3, indice_coluna, None, formatos["total_rotulo"])
    planilha.set_row(3, 20)
    for indice_coluna, cabecalho in enumerate(cabecalhos):
        largura = min(max(len(cabecalho) + 4, 16), 42)
        planilha.set_column(indice_coluna, indice_coluna, largura)
    for coluna_monetaria in ("E", "F", "G", "H", "I"):
        planilha.set_column(f"{coluna_monetaria}:{coluna_monetaria}", 20, formatos["numero"])

    cabecalhos_exibicao = [
        {
            "saldo_inicial_em_moeda_da_empresa": "saldo_inicial",
            "saldo_devedor_em_moeda_da_empresa": "saldo_devedor",
            "saldo_credor_em_moeda_da_empresa": "saldo_credor",
            "movimento_periodo": "movimento_periodo",
            "saldo_final_na_moeda_da_empresa": "saldo_final",
        }.get(cabecalho, cabecalho)
        for cabecalho in cabecalhos
    ]

    linha_cabecalho = 4
    for indice_linha, valores in enumerate(linhas, start=linha_cabecalho + 1):
        for indice_coluna, valor in enumerate(valores[: len(cabecalhos)]):
            if isinstance(valor, int | float):
                planilha.write_number(indice_linha, indice_coluna, arredondar_duas_casas(valor), formatos["numero"])
            else:
                planilha.write(indice_linha, indice_coluna, valor)

    if linhas:
        planilha.add_table(
            linha_cabecalho,
            0,
            linha_cabecalho + len(linhas),
            len(cabecalhos) - 1,
            {
                "name": "TabelaBalancete",
                "style": ESTILO_TABELA_PADRAO,
                "columns": colunas_tabela_padrao(cabecalhos_exibicao, formatos),
            },
        )


def gerar_resumo_soma_trimestral(
    caminho_sqlite: Path,
    pasta_saida: Path,
    balancete: tuple[list[str], list[list[object]]],
    contexto: dict[str, str],
) -> Path | None:
    # Consulta o SQLite consolidado e grava o resultado final em Excel.
    # Saida de controle: equivale a um SOMASE por conta contabil no Excel.
    # A soma ja considera a regra de inversao de sinal aplicada antes de gravar no SQLite.
    pasta_saida.mkdir(parents=True, exist_ok=True)
    empresa_arquivo = nome_seguro_arquivo(contexto["empresa"])
    arquivo_resumo = pasta_saida / f"{TIMESTAMP_RELATORIO}_empresa_{empresa_arquivo}_resumo_soma_trimestral_razoes.xlsx"

    with sqlite3.connect(caminho_sqlite) as conexao:
        # Consolida os resumos parciais de cada arquivo em uma visao trimestral por conta.
        linhas_resumo = conexao.execute(
            f"""
            SELECT
                {COLUNA_CODIGO_CONTA_EXTRACAO},
                {COLUNA_CODIGO_CONTA},
                {COLUNA_DESCRICAO_CONTA},
                ROUND(SUM({COLUNA_MONTANTE_MOEDA_EMPRESA}), 2) AS {COLUNA_SOMA_TRIMESTRAL}
            FROM {TABELA_RESUMO_ARQUIVO}
            GROUP BY
                {COLUNA_CODIGO_CONTA_EXTRACAO},
                {COLUNA_CODIGO_CONTA},
                {COLUNA_DESCRICAO_CONTA}
            ORDER BY {COLUNA_CODIGO_CONTA}
            """
        ).fetchall()
        linhas_dia = conexao.execute(
            f"""
            SELECT
                {COLUNA_NOME_ARQUIVO},
                {COLUNA_DATA_LANCTO_CONTABIL},
                {COLUNA_CONTA_RAZAO_DIARIA},
                ROUND(SUM({COLUNA_SALDOS_RAZOES}), 2) AS {COLUNA_SALDOS_RAZOES}
            FROM {TABELA_RESUMO_DIA}
            GROUP BY
                {COLUNA_NOME_ARQUIVO},
                {COLUNA_DATA_LANCTO_CONTABIL},
                {COLUNA_CONTA_RAZAO_DIARIA}
            ORDER BY
                {COLUNA_NOME_ARQUIVO},
                substr({COLUNA_DATA_LANCTO_CONTABIL}, 7, 4) || substr({COLUNA_DATA_LANCTO_CONTABIL}, 4, 2) || substr({COLUNA_DATA_LANCTO_CONTABIL}, 1, 2),
                {COLUNA_CONTA_RAZAO_DIARIA}
            """
        ).fetchall()

    if not linhas_resumo:
        print()
        print("Resumo trimestral nao gerado. O SQLite temporario nao possui dados para agrupar.")
        return None

    workbook = xlsxwriter.Workbook(arquivo_resumo)
    cabecalhos_balancete, linhas_balancete = balancete
    escrever_aba_menu(workbook, contexto)
    escrever_aba_conciliacao(workbook, linhas_resumo, cabecalhos_balancete, linhas_balancete, contexto)
    escrever_aba_balancete(workbook, cabecalhos_balancete, linhas_balancete, contexto)
    planilha = workbook.add_worksheet("Resumo - Razões")

    formatos = criar_formatos_relatorio(workbook)
    planilha.hide_gridlines(2)
    planilha.set_tab_color(COR_TITULO_RELATORIO)
    planilha.set_column("A:A", 36)
    planilha.set_column("B:B", 18)
    planilha.set_column("C:C", 32)
    planilha.set_column("D:D", 20, formatos["numero"])
    escrever_cabecalho_relatorio(planilha, formatos, 3, "Resumo da Soma Trimestral dos Razoes", contexto, True, 3)

    # Linha de totais do resumo trimestral: consolida a coluna monetaria dos razoes agrupados.
    total_soma_trimestral = arredondar_duas_casas(sum(numero_balancete(linha[3]) for linha in linhas_resumo))
    planilha.write("A4", "TOTAIS:", formatos["total_rotulo"])
    planilha.write_blank("B4", None, formatos["total_rotulo"])
    planilha.write_blank("C4", None, formatos["total_rotulo"])
    planilha.write_number("D4", total_soma_trimestral, formatos["total_numero"])
    planilha.set_row(3, 20)
    for linha, (codigo_conta_extracao, codigo_conta, descricao_conta, soma_trimestral) in enumerate(
        linhas_resumo, start=5
    ):
        planilha.write(linha, 0, codigo_conta_extracao)
        planilha.write(linha, 1, codigo_conta)
        planilha.write(linha, 2, descricao_conta)
        planilha.write_number(linha, 3, arredondar_duas_casas(soma_trimestral))

    planilha.add_table(
        4,
        0,
        len(linhas_resumo) + 4,
        3,
        {
            "name": "TabelaResumoSomaTrimestral",
            "style": ESTILO_TABELA_PADRAO,
            "columns": colunas_tabela_padrao(
                [COLUNA_CODIGO_CONTA_EXTRACAO, COLUNA_CODIGO_CONTA, COLUNA_DESCRICAO_CONTA, COLUNA_SOMA_TRIMESTRAL],
                formatos,
            ),
        },
    )

    escrever_aba_razoes_por_dia(workbook, linhas_dia, contexto)

    workbook.close()

    print()
    print("=" * 80)
    print("RESUMO TRIMESTRAL GERADO")
    print("=" * 80)
    print(f"Contas distintas : {len(linhas_resumo):,}")
    print(f"Arquivo Excel    : {arquivo_resumo}")
    print("=" * 80)

    return arquivo_resumo

def criar_parser() -> argparse.ArgumentParser:
    # Parser concentra os controles que antes exigiriam editar constantes no topo do arquivo.
    parser = argparse.ArgumentParser(
        description="Valida cabecalhos, le razoes SAP e gera resumo trimestral por conta."
    )
    parser.add_argument(
        "--modo",
        choices=sorted(MODOS_EXECUCAO),
        default=MODO_PADRAO,
        help="Perfil de execucao: teste, dev ou producao.",
    )
    parser.add_argument(
        "--pasta-bases",
        type=Path,
        help="Pasta com os arquivos .xlsx de razao. Se omitida, o usuario sera obrigado a selecionar a pasta.",
    )
    parser.add_argument(
        "--pasta-saida",
        type=Path,
        help="Pasta para salvar o resumo trimestral e o log. Se omitida, o usuario sera obrigado a selecionar a pasta.",
    )
    parser.add_argument(
        "--arquivo-balancete",
        type=Path,
        help="Arquivo .xlsx do balancete. Se omitido, o usuario sera obrigado a selecionar o arquivo.",
    )
    parser.add_argument(
        "--limite-arquivos",
        type=int,
        help="Limita a quantidade de arquivos processados, sobrescrevendo o modo escolhido.",
    )
    parser.add_argument(
        "--sem-janela",
        action="store_true",
        help="Nao abre seletores de pasta, mesmo nos modos dev/producao.",
    )
    parser.add_argument(
        "--parar-se-inconsistente",
        action="store_true",
        help="Interrompe antes da consolidacao se houver REEXTRAIR ou ERRO.",
    )
    return parser

def montar_configuracao(args: argparse.Namespace) -> ConfiguracaoExecucao:
    # Parte do perfil escolhido e aplica sobrescritas pontuais vindas da CLI.
    modo = MODOS_EXECUCAO[args.modo]
    limite_arquivos = args.limite_arquivos if args.limite_arquivos is not None else modo["limite_arquivos"]
    escolher_pasta_entrada = bool(modo["escolher_pasta_entrada"])
    escolher_pasta_saida = bool(modo["escolher_pasta_saida"])

    # Opcao para execucao automatizada: so e valida se a pasta de entrada vier por parametro.
    if args.sem_janela:
        escolher_pasta_entrada = False
        escolher_pasta_saida = False

    pausar_ao_final = not args.sem_janela

    return ConfiguracaoExecucao(
        modo=args.modo,
        pasta_bases=args.pasta_bases,
        pasta_saida_resumo=args.pasta_saida,
        arquivo_balancete=args.arquivo_balancete,
        limite_arquivos=limite_arquivos,
        escolher_pasta_entrada=escolher_pasta_entrada,
        escolher_pasta_saida=escolher_pasta_saida,
        abrir_relatorio=bool(modo["abrir_relatorio"]),
        parar_se_houver_inconsistencia=bool(modo["parar_se_houver_inconsistencia"])
        or args.parar_se_inconsistente,
        pausar_ao_final=pausar_ao_final,
    )

def main(configuracao: ConfiguracaoExecucao) -> Path | None:
    # Orquestracao principal: validacao, carga incremental, resumo e saidas ao usuario.
    print(f"Modo de execucao     : {configuracao.modo}")

    # A entrada e a saida sao resolvidas primeiro porque todas as etapas dependem dessas pastas.
    pasta_bases = obter_pasta_bases(configuracao.pasta_bases, configuracao.escolher_pasta_entrada)
    pasta_saida_resumo = obter_pasta_saida_resumo(
        configuracao.pasta_saida_resumo,
        configuracao.escolher_pasta_saida,
    )
    arquivo_balancete = obter_arquivo_balancete(
        configuracao.arquivo_balancete,
        configuracao.escolher_pasta_entrada,
    )
    cabecalhos_balancete, linhas_balancete = ler_balancete(arquivo_balancete)
    arquivo_relatorio = pasta_saida_resumo / "log" / f"{TIMESTAMP_RELATORIO}_validacao_cabecalhos_razoes.xlsx"
    arquivos = listar_arquivos_xlsx(pasta_bases, configuracao.limite_arquivos)

    inicio_validacao = perf_counter()
    mapeamento = validar_cabecalhos(arquivos, pasta_bases)
    tempo_validacao = perf_counter() - inicio_validacao

    if not mapeamento:
        return None

    salvar_relatorio_xlsx(mapeamento, pasta_bases, tempo_validacao, arquivo_relatorio)

    total, ok, reextrair, erro = contar_status(mapeamento)
    arquivos_reextrair = [item for item in mapeamento if item["status"] == "REEXTRAIR"]
    arquivos_erro = [item for item in mapeamento if item["status"] == "ERRO"]

    print()
    print("=" * 80)
    print("RESUMO DA VALIDACAO")
    print("=" * 80)
    print(f"Arquivos validados   : {total}")
    print(f"Arquivos OK          : {ok}")
    print(f"Reextrair do SAP     : {reextrair}")
    print(f"Erros de leitura     : {erro}")
    print(f"Tempo validacao      : {tempo_validacao:.2f} segundos")
    print(f"Relatorio Excel      : {arquivo_relatorio}")

    if arquivos_reextrair:
        print()
        print("Arquivos que precisam ser extraidos novamente do SAP:")

        for item in arquivos_reextrair:
            print(f"  - {item['arquivo']} | Campos faltando: {item['campos_faltando']}")

    if arquivos_erro:
        print()
        print("Arquivos com erro tecnico de leitura:")

        for item in arquivos_erro:
            print(f"  - {item['arquivo']} | Erro: {item['erro']}")

    abrir_relatorio_se_necessario(reextrair, erro, configuracao.abrir_relatorio, arquivo_relatorio)

    # Em producao, evita gerar consolidado com arquivos fora do layout esperado.
    if configuracao.parar_se_houver_inconsistencia and (reextrair > 0 or erro > 0):
        print()
        print("Processo interrompido: corrija/reextraia os arquivos apontados antes da consolidacao.")
        print("=" * 80)
        return None

    print("=" * 80)

    pasta_trabalho_sqlite = pasta_saida_resumo / "_sqlite_temp"
    caminho_sqlite = criar_sqlite_temporario(pasta_trabalho_sqlite)

    contexto_execucao = montar_contexto_execucao(arquivo_balancete, cabecalhos_balancete, linhas_balancete)

    try:
        carregar_resumos_no_sqlite(arquivos, caminho_sqlite)
        arquivo_resumo = gerar_resumo_soma_trimestral(
            caminho_sqlite,
            pasta_saida_resumo,
            (cabecalhos_balancete, linhas_balancete),
            contexto_execucao,
        )
    finally:
        # Limpeza explicita do banco e da subpasta operacional, quando estiver vazia.
        sqlite_apagado = apagar_sqlite_temporario(caminho_sqlite)
        if sqlite_apagado:
            try:
                pasta_trabalho_sqlite.rmdir()
            except OSError:
                pass

    abrir_excel_final_se_necessario(configuracao.abrir_relatorio, arquivo_resumo)

    return arquivo_resumo

def pausar_console_se_necessario(configuracao: ConfiguracaoExecucao) -> None:
    # Evita que a janela do console feche imediatamente no uso por duplo clique.
    # Em execucoes tecnicas com --sem-janela, nao pausa para preservar automacao.
    if not configuracao.pausar_ao_final:
        return

    try:
        print()
        input("Pressione ENTER para fechar esta janela...")
    except EOFError:
        pass

# Mantem o script importavel sem executar o fluxo automaticamente.
if __name__ == "__main__":
    argumentos = criar_parser().parse_args()
    configuracao = montar_configuracao(argumentos)

    try:
        arquivo_resumo = main(configuracao)

        if arquivo_resumo is not None:
            print()
            print(f"Processo finalizado. Resumo gerado em: {arquivo_resumo}")
    finally:
        pausar_console_se_necessario(configuracao)

