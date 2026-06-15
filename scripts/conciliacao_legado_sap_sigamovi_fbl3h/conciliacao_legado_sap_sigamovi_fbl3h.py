# =============================================================================
# Conciliacao Legado x SAP - Sigamovi x FBL3H
# Padrao seguro para Excel: sem Table, sem AutoFilter e sem celulas mescladas
# =============================================================================

from __future__ import annotations

import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# -----------------------------------------------------------------------------
# Configuracao
# -----------------------------------------------------------------------------

DEBUG = False
EXTENSOES_VALIDAS = {".xlsx", ".xls", ".csv"}
FMT_MONETARIO = '_-#,##0.00_-;[Red](#,##0.00)_-;_-* "-"??_-;_-@_-'

COR_AZUL_ESCURO = "1F3864"
COR_AZUL_MEDIO = "2F5496"
COR_CINZA_CLARO = "D9D9D9"
COR_BRANCO = "FFFFFF"
COR_VERDE = "E2EFDA"
COR_VERMELHO = "FFDADA"
COR_LARANJA = "FCE4D6"

COLUNAS_SIGAMOVI = {
    "SGMOVIEMPRESA",
    "SGMOVIFILIAL",
    "EMPRESA",
    "SGMOVIDEBITO",
    "SGMOVICREDITO",
    "SUM_of_SGMOVIVALOR",
    "SGMOVICCD",
    "SGMOVIROTINA",
    "SGMOVIGERADO",
    "IUNICODEMPRESA",
    "SGMOVANO",
    "SGMOVMES",
}

COLUNAS_FBL3H = {
    "Empresa",
    "Exercicio",
    "Periodo contabil",
    "Tipo de documento",
    "Conta do Razao",
    "Conta do Razao: texto breve",
    "Codigo da moeda empresa",
    "Valor em moeda da empresa",
    "Referencia",
    "Divisao",
}

# Variantes com acento podem aparecer em extracoes SAP.
ALIASES_FBL3H = {
    "Exercício": "Exercicio",
    "Período contábil": "Periodo contabil",
    "Conta do Razão": "Conta do Razao",
    "Código da moeda empresa": "Codigo da moeda empresa",
    "Referência": "Referencia",
    "Divisão": "Divisao",
}

COLUNAS_DEPARA_CONTA = {"MP_CONTA", "CONTA"}
COLUNAS_DEPARA_EMPRESA = {"MP_EMPRESA", "DIVISAO"}
ORDEM_ABAS = [
    "Resumo",
    "Resumo por Empresa",
    "Sigamovi",
    "Razao_SAP",
    "De_Para_Empresa_sas",
    "De_Para_Conta_sas",
]


@dataclass
class ArquivosConciliacao:
    legado: Path
    sap: Path
    depara_conta: Path
    depara_empresa: Path
    sap_correcao: Path | None = None


def selecionar_pasta(titulo: str, mensagem: str, initialdir: str | None = None) -> Path:
    root = tk.Tk()
    root.withdraw()
    root.lift()
    root.attributes("-topmost", True)
    messagebox.showinfo("Selecao de Pasta", mensagem)
    pasta = filedialog.askdirectory(title=titulo, initialdir=initialdir)
    root.destroy()
    if not pasta:
        alertar_e_encerrar("Processo Encerrado", f"Nenhuma pasta selecionada para: {titulo}")
    return Path(pasta)


def alertar_e_encerrar(titulo: str, mensagem: str) -> None:
    root = tk.Tk()
    root.withdraw()
    root.lift()
    root.attributes("-topmost", True)
    messagebox.showwarning(titulo, mensagem)
    root.destroy()
    print(f"\n[ENCERRADO] {titulo}: {mensagem}")
    sys.exit(0)


def listar_arquivos_validos(pasta: Path) -> list[Path]:
    arquivos = [
        p for p in pasta.iterdir()
        if p.is_file() and p.suffix.lower() in EXTENSOES_VALIDAS and not p.name.startswith("~$")
    ]
    if not arquivos:
        alertar_e_encerrar(
            "Processo Encerrado",
            f"A pasta {pasta.name} nao contem arquivos Excel ou CSV validos.",
        )
    return sorted(arquivos)


def ler_amostra(caminho: Path, nrows: int = 100) -> pd.DataFrame | None:
    try:
        if caminho.suffix.lower() == ".csv":
            return pd.read_csv(caminho, nrows=nrows, sep=None, engine="python")
        return pd.read_excel(caminho, nrows=nrows)
    except Exception as exc:
        if DEBUG:
            print(f"Falha ao ler amostra {caminho.name}: {exc}")
        return None


def normalizar_colunas_fbl3h(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns={k: v for k, v in ALIASES_FBL3H.items() if k in df.columns})


def faltantes(df: pd.DataFrame, colunas: set[str]) -> set[str]:
    return colunas - set(df.columns)


def identificar_arquivos(arquivos: list[Path]) -> ArquivosConciliacao:
    amostras: dict[Path, pd.DataFrame] = {}
    usados: set[Path] = set()
    sap = legado = depara_conta = depara_empresa = sap_correcao = None

    for arq in arquivos:
        df = ler_amostra(arq)
        if df is not None:
            amostras[arq] = df

    # FBL3H principal
    for arq, df0 in amostras.items():
        df = normalizar_colunas_fbl3h(df0)
        if faltantes(df, COLUNAS_FBL3H):
            continue
        tipo_ok = df["Tipo de documento"].astype(str).eq("OI").any()
        ref = df["Referencia"].astype(str)
        ref_ok = ref.str.startswith("580000" + chr(124) + "V").any()
        if tipo_ok and ref_ok:
            sap = arq
            usados.add(arq)
            break

    # Sigamovi
    for arq, df in amostras.items():
        if arq in usados:
            continue
        if not faltantes(df, COLUNAS_SIGAMOVI):
            legado = arq
            usados.add(arq)
            break

    # De-para conta
    for arq, df in amostras.items():
        if arq in usados:
            continue
        if not faltantes(df, COLUNAS_DEPARA_CONTA):
            depara_conta = arq
            usados.add(arq)
            break

    # De-para empresa
    for arq, df in amostras.items():
        if arq in usados:
            continue
        if not faltantes(df, COLUNAS_DEPARA_EMPRESA):
            depara_empresa = arq
            usados.add(arq)
            break

    # FBL3H correcao opcional
    for arq, df0 in amostras.items():
        if arq in usados:
            continue
        df = normalizar_colunas_fbl3h(df0)
        if faltantes(df, COLUNAS_FBL3H):
            continue
        ref = df["Referencia"].astype(str)
        tipo = df["Tipo de documento"].astype(str)
        ref_corr = ref.str.startswith("580000" + chr(124) + "V").any()
        nao_oi = ~tipo.eq("OI").all()
        if ref_corr and nao_oi:
            sap_correcao = arq
            usados.add(arq)
            break

    pendencias = []
    if legado is None:
        pendencias.append("Sigamovi")
    if sap is None:
        pendencias.append("FBL3H principal")
    if depara_conta is None:
        pendencias.append("De-Para Conta")
    if depara_empresa is None:
        pendencias.append("De-Para Empresa")

    if pendencias:
        alertar_e_encerrar(
            "Arquivos obrigatorios nao identificados",
            "Arquivos pendentes: " + ", ".join(pendencias),
        )

    return ArquivosConciliacao(
        legado=legado,
        sap=sap,
        depara_conta=depara_conta,
        depara_empresa=depara_empresa,
        sap_correcao=sap_correcao,
    )


def ler_excel_ou_csv(caminho: Path, **kwargs) -> pd.DataFrame:
    if caminho.suffix.lower() == ".csv":
        return pd.read_csv(caminho, sep=None, engine="python", **kwargs)
    return pd.read_excel(caminho, **kwargs)


def normalizar_codigo_serie(s: pd.Series) -> pd.Series:
    return (
        s.astype(str)
        .str.replace(".0", "", regex=False)
        .str.strip()
        .replace({"nan": "", "None": ""})
    )


def preparar_bases(arqs: ArquivosConciliacao):
    depara_conta = ler_excel_ou_csv(arqs.depara_conta, dtype=str)
    depara_conta["MP_CONTA"] = normalizar_codigo_serie(depara_conta["MP_CONTA"])
    depara_conta["CONTA"] = normalizar_codigo_serie(depara_conta["CONTA"])

    depara_empresa = ler_excel_ou_csv(arqs.depara_empresa, dtype=str)
    depara_empresa["MP_EMPRESA"] = normalizar_codigo_serie(depara_empresa["MP_EMPRESA"])
    depara_empresa["DIVISAO"] = normalizar_codigo_serie(depara_empresa["DIVISAO"])

    sigamovi_raw = ler_excel_ou_csv(
        arqs.legado,
        dtype={"EMPRESA": str, "SGMOVIDEBITO": str, "SGMOVICREDITO": str, "SGMOVIGERADO": str},
    )
    sigamovi = sigamovi_raw[sigamovi_raw["SGMOVIGERADO"].astype(str).str.strip() == "S"].copy()
    sigamovi["EMPRESA"] = normalizar_codigo_serie(sigamovi["EMPRESA"])
    sigamovi["SGMOVIDEBITO"] = normalizar_codigo_serie(sigamovi["SGMOVIDEBITO"])
    sigamovi["SGMOVICREDITO"] = normalizar_codigo_serie(sigamovi["SGMOVICREDITO"])
    sigamovi["SUM_of_SGMOVIVALOR"] = pd.to_numeric(sigamovi["SUM_of_SGMOVIVALOR"], errors="coerce").fillna(0)

    fbl3h_raw = normalizar_colunas_fbl3h(ler_excel_ou_csv(arqs.sap))
    fbl3h = preparar_fbl3h(fbl3h_raw, depara_conta)

    fbl3h_correcao_raw = None
    fbl3h_correcao = None
    if arqs.sap_correcao:
        fbl3h_correcao_raw = normalizar_colunas_fbl3h(ler_excel_ou_csv(arqs.sap_correcao))
        fbl3h_correcao = preparar_fbl3h(fbl3h_correcao_raw, depara_conta)

    return depara_conta, depara_empresa, sigamovi_raw, sigamovi, fbl3h_raw, fbl3h, fbl3h_correcao_raw, fbl3h_correcao


def preparar_fbl3h(df: pd.DataFrame, depara_conta: pd.DataFrame) -> pd.DataFrame:
    base = df.dropna(subset=["Empresa"]).copy()
    mapa_conta = depara_conta.set_index("CONTA")["MP_CONTA"].to_dict()
    base["Divisao"] = normalizar_codigo_serie(base["Divisao"])
    base["CONTA_SAP"] = normalizar_codigo_serie(base["Conta do Razao"])
    base["VALOR_SAP"] = pd.to_numeric(base["Valor em moeda da empresa"], errors="coerce").fillna(0)
    base["MP_CONTA"] = base["CONTA_SAP"].map(mapa_conta)
    return base


def explodir_debito_credito(sigamovi: pd.DataFrame) -> pd.DataFrame:
    deb = sigamovi[["EMPRESA", "SGMOVIDEBITO", "SUM_of_SGMOVIVALOR"]].copy()
    deb.columns = ["MP_EMPRESA", "MP_CONTA", "VALOR"]
    deb["SINAL"] = 1

    cred = sigamovi[["EMPRESA", "SGMOVICREDITO", "SUM_of_SGMOVIVALOR"]].copy()
    cred.columns = ["MP_EMPRESA", "MP_CONTA", "VALOR"]
    cred["SINAL"] = -1

    movs = pd.concat([deb, cred], ignore_index=True)
    movs["MP_EMPRESA"] = normalizar_codigo_serie(movs["MP_EMPRESA"])
    movs["MP_CONTA"] = normalizar_codigo_serie(movs["MP_CONTA"])
    movs["VALOR_LIQ"] = pd.to_numeric(movs["VALOR"], errors="coerce").fillna(0) * movs["SINAL"]
    return movs


def conciliar(movs: pd.DataFrame, depara_conta: pd.DataFrame, depara_empresa: pd.DataFrame, fbl3h: pd.DataFrame, fbl3h_correcao: pd.DataFrame | None) -> tuple[pd.DataFrame, pd.DataFrame]:
    contas_escopo = set(depara_conta["MP_CONTA"].dropna())
    mapa_empresa = depara_empresa.set_index("MP_EMPRESA")["DIVISAO"].to_dict()

    movs_escopo = movs[movs["MP_CONTA"].isin(contas_escopo)].copy()
    movs_escopo["DIVISAO"] = movs_escopo["MP_EMPRESA"].map(mapa_empresa)

    saldo_legado = (
        movs_escopo[movs_escopo["DIVISAO"].notna()]
        .groupby(["DIVISAO", "MP_CONTA"], as_index=False)["VALOR_LIQ"]
        .sum()
        .rename(columns={"VALOR_LIQ": "SALDO_LEGADO"})
    )

    fbl3h_concil = fbl3h[fbl3h["MP_CONTA"].notna()].copy()
    if fbl3h_correcao is not None:
        fbl3h_concil = pd.concat([fbl3h_concil, fbl3h_correcao[fbl3h_correcao["MP_CONTA"].notna()]], ignore_index=True)

    saldo_sap = (
        fbl3h_concil
        .groupby(["Divisao", "MP_CONTA"], as_index=False)["VALOR_SAP"]
        .sum()
        .rename(columns={"Divisao": "DIVISAO", "VALOR_SAP": "SALDO_SAP"})
    )

    conciliacao = pd.merge(saldo_legado, saldo_sap, on=["DIVISAO", "MP_CONTA"], how="outer")
    conciliacao["SALDO_LEGADO"] = conciliacao["SALDO_LEGADO"].fillna(0).round(2)
    conciliacao["SALDO_SAP"] = conciliacao["SALDO_SAP"].fillna(0).round(2)
    conciliacao["DIFERENCA"] = (conciliacao["SALDO_SAP"] - conciliacao["SALDO_LEGADO"]).round(2)
    conciliacao["STATUS"] = np.where(conciliacao["DIFERENCA"].abs() < 0.01, "OK", "DIVERGENTE")
    conciliacao["ORIGEM"] = np.select(
        [
            (conciliacao["SALDO_SAP"] == 0) & (conciliacao["SALDO_LEGADO"] != 0),
            (conciliacao["SALDO_LEGADO"] == 0) & (conciliacao["SALDO_SAP"] != 0),
        ],
        ["APENAS_LEGADO", "APENAS_SAP"],
        default="AMBOS",
    )
    conciliacao = conciliacao.sort_values(["STATUS", "DIVISAO", "MP_CONTA"]).reset_index(drop=True)

    resumo = conciliacao.groupby("MP_CONTA", as_index=False).agg(
        SALDO_LEGADO=("SALDO_LEGADO", "sum"),
        SALDO_SAP=("SALDO_SAP", "sum"),
    )
    resumo["DIFERENCA"] = (resumo["SALDO_SAP"] - resumo["SALDO_LEGADO"]).round(2)
    resumo["STATUS"] = np.where(resumo["DIFERENCA"].abs() < 0.01, "OK", "DIVERGENTE")
    resumo = resumo.sort_values(["STATUS", "MP_CONTA"]).reset_index(drop=True)
    return conciliacao, resumo


def escrever_excel(caminho: Path, abas: dict[str, pd.DataFrame], resumo_info: dict[str, str]) -> None:
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        for nome in ORDEM_ABAS:
            df = abas[nome]
            df.to_excel(writer, sheet_name=nome, index=False, startrow=5)

    wb = load_workbook(caminho)
    for nome in ORDEM_ABAS:
        ws = wb[nome]
        df = abas[nome]
        ncols = max(len(df.columns), 1)
        nrows = len(df)
        titulo = resumo_info.get(nome, nome)

        ws.cell(row=1, column=1, value=titulo)
        ws.cell(row=2, column=1, value="Evidencia operacional de conciliacao")
        ws.cell(row=3, column=1, value=f"Gerado em: {datetime.now():%d/%m/%Y %H:%M:%S}")
        ws.cell(row=4, column=1, value=f"Registros: {nrows}")
        ws.cell(row=5, column=1, value="Tabela a partir da linha 6")

        for row in range(1, 5):
            for col in range(1, ncols + 1):
                c = ws.cell(row=row, column=col)
                c.fill = PatternFill("solid", fgColor=COR_AZUL_ESCURO if row == 1 else COR_AZUL_MEDIO)
                c.font = Font(color=COR_BRANCO, bold=(row == 1), size=12 if row == 1 else 10)
                c.alignment = Alignment(vertical="center")
        for col in range(1, ncols + 1):
            c = ws.cell(row=6, column=col)
            c.fill = PatternFill("solid", fgColor=COR_AZUL_ESCURO)
            c.font = Font(color=COR_BRANCO, bold=True)
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = min(max(14, len(str(c.value)) + 2), 35)

        ws.freeze_panes = "A7"

        for idx, col_name in enumerate(df.columns, start=1):
            if any(chave in str(col_name).upper() for chave in ["SALDO", "VALOR", "DIFERENCA"]):
                for row in range(7, 7 + nrows):
                    ws.cell(row=row, column=idx).number_format = FMT_MONETARIO

        if nome in {"Resumo", "Resumo por Empresa"} and "STATUS" in df.columns:
            col_status = list(df.columns).index("STATUS") + 1
            for row in range(7, 7 + nrows):
                status = ws.cell(row=row, column=col_status).value
                fill = COR_VERDE if status == "OK" else COR_VERMELHO
                for col in range(1, ncols + 1):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill)

    wb._sheets = [wb[nome] for nome in ORDEM_ABAS if nome in wb.sheetnames]
    wb.save(caminho)


def main() -> None:
    print("\n" + "=" * 70)
    print("Conciliacao Legado x SAP - Sigamovi x FBL3H")
    print("=" * 70)

    input_dir = selecionar_pasta(
        "Bases extraidas do Legado x SAP",
        "Selecione a pasta com Sigamovi, FBL3H e de-paras.",
    )
    output_dir = selecionar_pasta(
        "Resultado da Conciliacao",
        "Selecione a pasta onde sera salva a evidencia em Excel.",
        initialdir=str(input_dir.parent),
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    arquivos = listar_arquivos_validos(input_dir)
    if len(arquivos) < 4:
        alertar_e_encerrar("Poucos arquivos", "Sao esperados ao menos 4 arquivos de entrada.")

    arqs = identificar_arquivos(arquivos)
    print(f"Legado: {arqs.legado.name}")
    print(f"SAP principal: {arqs.sap.name}")
    print(f"SAP correcao: {arqs.sap_correcao.name if arqs.sap_correcao else 'nao localizado'}")
    print(f"De-para conta: {arqs.depara_conta.name}")
    print(f"De-para empresa: {arqs.depara_empresa.name}")

    depara_conta, depara_empresa, sigamovi_raw, sigamovi, fbl3h_raw, fbl3h, fbl3h_correcao_raw, fbl3h_correcao = preparar_bases(arqs)
    movs = explodir_debito_credito(sigamovi)
    conciliacao, resumo = conciliar(movs, depara_conta, depara_empresa, fbl3h, fbl3h_correcao)

    razao_sap = fbl3h_raw.copy()
    razao_sap["ORIGEM"] = "integracao"
    if fbl3h_correcao_raw is not None:
        corr = fbl3h_correcao_raw.copy()
        corr["ORIGEM"] = "correcao"
        razao_sap = pd.concat([razao_sap, corr], ignore_index=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = output_dir / f"conciliacao_legado_sap_sigamovi_fbl3h_{ts}.xlsx"

    abas = {
        "Resumo": resumo,
        "Resumo por Empresa": conciliacao,
        "Sigamovi": sigamovi_raw,
        "Razao_SAP": razao_sap,
        "De_Para_Empresa_sas": depara_empresa,
        "De_Para_Conta_sas": depara_conta,
    }
    resumo_info = {
        "Resumo": "Resumo por Conta",
        "Resumo por Empresa": "Resumo por Divisao x Conta",
        "Sigamovi": "Base Sigamovi",
        "Razao_SAP": "Razao SAP FBL3H",
        "De_Para_Empresa_sas": "De-Para Empresas",
        "De_Para_Conta_sas": "De-Para Contas",
    }
    escrever_excel(output, abas, resumo_info)

    total = len(conciliacao)
    ok = int((conciliacao["STATUS"] == "OK").sum())
    divergente = total - ok
    print("\nResultado")
    print(f"Total combinacoes: {total}")
    print(f"OK: {ok}")
    print(f"Divergente: {divergente}")
    print(f"Arquivo gerado: {output}")


if __name__ == "__main__":
    main()
